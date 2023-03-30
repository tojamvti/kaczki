
import openpyxl
from openpyxl import load_workbook


def find_max_height(num_ducks, max_width, ducks):
    # Stworzenie tabeli dwuwymiarowej o wymiarach (liczba kaczuszek + 1) na (maksymalna szerokość rzędu + 1)
    table = [[0 for _ in range(max_width + 1)] for _ in range(num_ducks + 1)]

    # Wypełnienie tabeli
    for i in range(1, num_ducks + 1):
        for j in range(1, max_width + 1):
            # Kaczuszka nie może być dodana do rzędu
            if ducks[i-1][1] > j:
                table[i][j] = table[i-1][j]
            # Możliwe dodanie kaczuszki, wybór lepszego rozwiązania
            else:
                table[i][j] = max(table[i-1][j], table[i-1][j-ducks[i-1][1]] + ducks[i-1][0])

    # Odtworzenie rozwiązania (kaczuszki wybrane do ułożenia)
    result = []
    j = max_width
    for i in range(num_ducks, 0, -1):
        if table[i][j] != table[i-1][j]:
            result.append(i-1)
            j -= ducks[i-1][1]

    # Odwrócenie listy wynikowej (kaczuszki były dodawane od końca)
    

    return f"Maksymalna dostępna suma wysokości: {table[num_ducks][max_width]}"
    pass
def MAXIMUM_WIDTH():

    heights = []
    widths = []
    
    wb = openpyxl.load_workbook('kaczki.xlsx')
    ws = wb['Arkusz1']
    rows = ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2)

    for a, b in rows:
        heights.append(a.value)
        widths.append(b.value)

    chickens = list(zip(heights, widths))
    sorted_chickens = sorted(chickens, key=lambda x: x[0], reverse=True)
    sorted_chickens.pop(0)
    kurczak = chickens.pop(0)
    MAXIMUM_WIDTH = kurczak[1]

    return MAXIMUM_WIDTH
def ducks():
    heights = []
    widths = []
    
    wb = openpyxl.load_workbook('kaczki.xlsx')
    ws = wb['Arkusz1']
    rows = ws.iter_rows(min_row=1, max_row=21, min_col=1, max_col=2)

    for a, b in rows:
        heights.append(a.value)
        widths.append(b.value)

    chickens = list(zip(heights, widths))
    sorted_chickens = sorted(chickens, key=lambda x: x[0], reverse=True)
    sorted_chickens.pop(0)

    return sorted_chickens
def NR_OF_DUCKS():

    heights = []
    widths = []
    
    wb = openpyxl.load_workbook('kaczki.xlsx')
    ws = wb['Arkusz1']
    rows = ws.iter_rows(min_row=1, max_row=3, min_col=1, max_col=2)

    for a, b in rows:
        heights.append(a.value)
        widths.append(b.value)

    chickens = list(zip(heights, widths))
    sorted_chickens = sorted(chickens, key=lambda x: x[0], reverse=True)
    sorted_chickens.pop(0)
    kurczak = chickens.pop(0)
    NR_OF_DUCKS = kurczak[0]

    return NR_OF_DUCKS  

def kaczki():
    return (find_max_height(NR_OF_DUCKS(),MAXIMUM_WIDTH(),ducks()))

print(kaczki())
