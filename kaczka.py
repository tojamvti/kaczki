import openpyxl
from openpyxl import load_workbook

def ducks():
    
    heights = []
    widths = []
    
    wb = openpyxl.load_workbook('zadanie-rekrutacyjne.xlsx')
    ws = wb['Arkusz1'] 
    rows = ws.iter_rows(min_row=1,max_row=21,min_col=1,max_col=2) 

    for a,b in rows:
        heights.append(a.value)
        widths.append(b.value)
    
    MAXIMUM_WIDTH = widths[0]
    
    heights.pop(0)
    heights.sort(reverse=True)
    
    widths.pop(0)
    widths.sort(reverse=True)
   
    for i in heights:
        rediheight = 0
        rediwidth = 0
        while rediwidth <= MAXIMUM_WIDTH:
            rediheight += heights[i]
            rediwidth += widths[i]
            
    return f"maksymalna dostępna suma: {rediheight}"

print(ducks())
